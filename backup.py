from assemble_rules_update_inputs_file import my_assemble_dict
import os, glob
import datetime
from subprocess import check_output as qx
import subprocess
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill

class AssembleAutomation:

    def __init__(self):
        self.rowNum = 1
        self.new_assemble_rules_path = ""
        self.i = 1
        self.modified_assemble_rule_data = ""
        self.start_time = datetime.datetime.now().time().strftime('%H:%M:%S')
        self.thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

    def replaceKeys(self):
        folder = my_assemble_dict["$Rules_Path"]

        # we are storing the rule after update into a new folder. so we are making sure that folder is exist.
        if not os.path.exists(my_assemble_dict["$Rules_modified_Path"]):
            self.new_assemble_rules_path = os.makedirs(my_assemble_dict["$Rules_modified_Path"])
        else :
            self.new_assemble_rules_path = my_assemble_dict["$Rules_modified_Path"]

        # if reports folder is not exist, then create it now.
        if not os.path.exists(my_assemble_dict["$Assemble_Logs_Report_Path"]):
            Assemble_Logs_Report_Path = os.makedirs(my_assemble_dict["$Assemble_Logs_Report_Path"])
        else :
            Assemble_Logs_Report_Path = my_assemble_dict["$Assemble_Logs_Report_Path"]

        print "List of rules going to be Executed :\n" + "====================================="
        # getting all the files in folder and creating a proper path to access file.
        for file in os.listdir(folder):
            newfilePath = os.path.join(str(self.new_assemble_rules_path) + "/", file)
            # check for only .ini files
            if file.endswith(".ini"):

                print str(self.i) + ") " + file
                self.i += 1

                filepath = os.path.join(folder, file)
                # opening each ini file with keys present.
                with open(filepath, 'r') as main:
                    input_data = main.read()
                    main.close()
                    # writing the .ini data to a new .ini file.
                    with open(newfilePath, "w") as new_main:
                        # iterating over all the keys and replace with the values in the .ini file.
                        for key, value in my_assemble_dict.iteritems():
                            # if reports path is not available, then we are creating and appending it here.
                            if key == "$Assemble_Logs_Report_Path":
                                my_assemble_dict["$Assemble_Logs_Report_Path"] = Assemble_Logs_Report_Path
                            if key == "$SendEmailSubject":
                                value = file.strip(".ini")
                            elif key == "$SendEmailBody":
                                value = file.strip(".ini")
                            elif key == "$ReportMessage":
                                value = file.strip(".ini")
                            elif key == "$SendMessageText":
                                value = file.strip(".ini")
                            elif key == "$ActionReason":
                                value = file.strip(".ini")
                            elif key == "$ReportName":
                                value = file.strip(".ini")
                            input_data = input_data.replace(key, value)
                            self.modified_assemble_rule_data = input_data
                        # writing the real data that can be used for execute as a rule into a new ini file.
                        new_main.write(self.modified_assemble_rule_data)

    def ExecuteFiles(self):
        # Create a folder to execute all the rules and run them in the same path.
        if not os.path.exists(my_assemble_dict["$AssembleExecutionLogsPath"]):
            Logs_Assemble_run = os.makedirs(my_assemble_dict["$AssembleExecutionLogsPath"])
        else:
            Logs_Assemble_run = my_assemble_dict["$AssembleExecutionLogsPath"]

        # change the working dir to logs folder. so that all Assemble run logs will be saved here.
        os.chdir(Logs_Assemble_run)
        print "\nPresent Working Directory :", os.getcwd()
        self.i = 1

       # self.new_assemble_rules_path = "C:\Users\Miadmin\Desktop\Assemble_bug_verification\/assemble_v47.1\/rules\/"
        for file in os.listdir(self.new_assemble_rules_path):
            if file.endswith("ini"):
                # windows have system generated file named desktop.ini. we are eliminating it being executed.
                # we cannot delete this file manually. it will be created automatically even if we remove it.
                if file.startswith("desktop"):
                    continue
                reportFile = file
                lenStr = "\n" + str(self.i) + ") " + file
                print lenStr
                self.i += 1
                print "#" * len(lenStr)
                file = self.new_assemble_rules_path + "\/" + file
                args = my_assemble_dict["$AssembleExeLocation"] + " " + my_assemble_dict[
                    "$VSPiniLocation"] + " " + file + " " + my_assemble_dict["$ModeOfRuleExecution"]
                print "\n=============================================================="
                print "Rule Executing :\n", args
                try:
                    FinalOutput = qx(args)
                    print "\nOutput of the Rule :\n", FinalOutput
                    if FinalOutput.endswith("1"):
                        results = "PASS"
                        print "\nFinalResult of the Rule : ", "PASS"
                    if FinalOutput.endswith("0"):
                        results = "FAIL"
                        print "\nFinalResult of the Rule : ", "FAIL"

                    # the assemble log for the above rule execution
                    list_of_files = glob.glob(Logs_Assemble_run + "*.log")
                    latest_file = max(list_of_files, key=os.path.getctime)

                    self.write_results_to_workbook(reportFile, results, latest_file)
                    print "==============================================================\n"
                except subprocess.CalledProcessError as e:
                    print e.message
                    results = "ERROR"
                    # the assemble log for the above rule execution
                    list_of_files = glob.glob(Logs_Assemble_run + "*.log")
                    latest_file = max(list_of_files, key=os.path.getctime)
                    self.write_results_to_workbook(reportFile, results, latest_file)
                    continue

    def write_results_to_workbook(self, reportFile, FinalOutput, latest_file):
        filePath = my_assemble_dict["$Rules_Path"] + "\/" + 'Assemble_Run_TestResults.xlsx'
        if not os.path.isfile(filePath):
            wb = Workbook()
            worksheet1 = wb.worksheets[0]
        else:
            wb = load_workbook(filePath)
            worksheet1 = wb.worksheets[0]

        if self.rowNum == 1:
            # Write some data headers.
            worksheet1.cell('A1').border = self.thin_border
            worksheet1.cell('A1').fill = PatternFill(bgColor="FFC7CE", fill_type = "solid")
            worksheet1['A1'] = 'iniFile'

            worksheet1.cell('B1').border = self.thin_border
            worksheet1.cell('B1').fill = PatternFill(bgColor="FFC7CE", fill_type="solid")
            worksheet1['B1'] = 'Result'

            worksheet1.cell('C1').border = self.thin_border
            worksheet1.cell('C1').fill = PatternFill(bgColor="FFC7CE", fill_type="solid")
            worksheet1['C1'] = 'LogLocation'

        self.rowNum += 1
        # Write some data from row 2 onwards.
        worksheet1.cell('A'+str(self.rowNum), column=0).value = reportFile
        worksheet1.cell('A'+str(self.rowNum), column=0).border = self.thin_border
        worksheet1.cell('B'+str(self.rowNum), column=1).value = FinalOutput
        worksheet1.cell('B'+str(self.rowNum), column=1).border = self.thin_border
        worksheet1.cell('C'+str(self.rowNum), column=2).hyperlink = latest_file
        worksheet1.cell('C'+str(self.rowNum), column=2).border = self.thin_border

        wb.save(filePath)
        wb.close()

    def roundtrip(self):
        self.end_time = datetime.datetime.now().time().strftime('%H:%M:%S')
        self.totalroundtrip = (datetime.datetime.strptime(self.end_time, '%H:%M:%S') - datetime.datetime.strptime(self.start_time, '%H:%M:%S'))
        print "@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@"
        print "\n Test was Started on : ", self.start_time
        print "\n Test is Ended on : ", self.end_time
        print "\nTotal time taken to complete the Run : ", self.totalroundtrip
        print "\n      - Days : ", self.totalroundtrip.days
        print "\n      - Days : ", self.totalroundtrip.seconds
        print "@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@"

    def decider(self):
        print "This Script is created by providing ability to execute all the scripts at a time or the script given by you."
        print " - Press 1 to execute all the runs!"
        print " - Press 2 to execute the given script!"
        choice1 = str(raw_input("Which one do you want to proceed with : "))
        trails = 1
        while not ((choice1 == "1") or (choice1 == "2")):
            print "Not a valid number. Please try again...!"
            choice1 = str(raw_input("Which one do you want to proceed with : "))
            if trails == 3:
                print "Exiting the Execution......!"
                break
            trails += 1
        return choice1

    def SingleScriptExecutor(self):
        assemble_executable_path = raw_input("Enter your Assemble Executable path : ")
        trails = 1
        while not str(assemble_executable_path).endswith(".exe"):
            assemble_executable_path = str(raw_input("Please Enter your Proper Assemble Executable path : "))
            if trails == 3:
                print "Exiting the Execution......!"
                exit()
            trails += 1

        vsp_path = raw_input("Please Enter your VSP.ini file path : ")
        trails = 1
        while not str(vsp_path).endswith(".ini"):
            vsp_path = str(raw_input("Please Enter your Proper VSP.ini file path : "))
            if trails == 3:
                print "Exiting the Execution......!"
                exit()
            trails += 1

        Assemble_rule = raw_input("Please Enter your Assemble rule file path : ")
        trails = 1
        while not str(Assemble_rule).endswith(".ini"):
            vsp_path = str(raw_input("Please Enter your Proper Assemble rule file path : "))
            if trails == 3:
                print "Exiting the Execution......!"
                exit()
            trails += 1

        test_mode = str(raw_input("Please Enter your test Execution Mode : "))
        trails = 1
        while not ((test_mode == "test") or (test_mode == "run")):
            vsp_path = str(raw_input("Please Enter your Proper Execution Mode (test/run) : "))
            if trails == 3:
                print "Exiting the Execution......!"
                exit()
            trails += 1

        Logs_path = str(raw_input("Where do you want to store the Logs : "))
        trails = 1
        while not ((os.path.exists(Logs_path))):
            vsp_path = str(raw_input("Path Doest not Exist. Please enter the correct path : "))
            if trails == 3:
                print "Exiting the Execution......!"
                exit()
            trails += 1

        os.chdir(Logs_path)
        ExecuteTheRule = qx(assemble_executable_path + " " + vsp_path + " " + Assemble_rule + " " + test_mode)
        print "\nOutput of the Rule :\n", ExecuteTheRule
        if ExecuteTheRule.endswith("1"):
            results = "PASS"
            print "\nFinalResult of the Rule : ", "PASS"
        if ExecuteTheRule.endswith("0"):
            results = "FAIL"
            print "\nFinalResult of the Rule : ", "FAIL"

if __name__ == "__main__":
    executionStart = AssembleAutomation()
    choice = executionStart.decider()
    if choice == "1":
        turn = raw_input("Are you sure want to proceed with all. Please press Enter..... ")
        trails = 1
        while not turn in " ":
            turn = raw_input("Please press Enter only.. : ")
            if trails == 3:
                print "Exiting the Execution......!"
                exit()
            trails += 1
        executionStart.ExecuteFiles()
        executionStart.roundtrip()
    elif choice == "2":
        executionStart.SingleScriptExecutor()